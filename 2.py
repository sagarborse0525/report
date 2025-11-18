#!/usr/bin/env python3
"""
Generate GitLab vulnerabilities report by scrum.

Outputs:
  - One sheet per scrum (project-level rows)
  - A Summary sheet in the first position with totals
  - A "percent change" table comparing window counts (30/60/90 days) to current open counts
    (format: +12.34%, -45.67%, or '-' when baseline current == 0 and window > 0)
"""

import os
import time
import logging
from typing import Dict, List, Optional, Tuple
import requests
import pandas as pd
from dateutil import parser
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta, timezone
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ---------------------- User-provided mapping ---------------------- #
scrum_ids = {
    "sme-mobile": "9715323",
    "sme-online": "101745074",
    "sme-finacle": "12225673",
    "sme-report": "115417730"
}

# ---------------------- Configuration ---------------------- #
GITLAB_ACCESS_TOKEN = os.environ.get("GRAPHQL_API_TOKEN")  # PAT/OAuth token
BASE_URL = "https://gitlab.com/api/v4"
OUTPUT_XLSX = "gitlab_vuln.xlsx"

# Tune network behavior
DEFAULT_PER_PAGE = 50
DEFAULT_TIMEOUT: Tuple[int, int] = (10, 60)  # (connect, read)
MAX_RETRIES = 5
BACKOFF_FACTOR = 0.5

# Logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)
logger = logging.getLogger("gitlab-vuln-report")


# ---------------------- HTTP Session with Retries ---------------------- #
def make_session(token: str) -> requests.Session:
    if not token:
        raise RuntimeError("GRAPHQL_API_TOKEN environment variable is required for authentication.")
    retry = Retry(
        total=MAX_RETRIES,
        connect=MAX_RETRIES,
        read=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods={"GET", "HEAD", "OPTIONS"},
        respect_retry_after_header=True,
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=50)
    s = requests.Session()
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    s.trust_env = True
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Accept": "application/json",
        "Connection": "keep-alive",
    })
    return s


SESSION = make_session(GITLAB_ACCESS_TOKEN)


# ---------------------- Helpers ---------------------- #
def _sleep_if_rate_limited(resp: requests.Response) -> None:
    if resp.status_code == 429:
        retry_after = resp.headers.get("Retry-After")
        if retry_after:
            try:
                delay = int(retry_after)
                logger.warning("Rate limited. Sleeping for %s seconds...", delay)
                time.sleep(delay)
            except ValueError:
                logger.warning("Rate limited without numeric Retry-After; backing off 2 seconds.")
                time.sleep(2)


def _get_json(url: str, params: Dict) -> Optional[List[Dict]]:
    try:
        resp = SESSION.get(url, params=params, timeout=DEFAULT_TIMEOUT)
    except requests.exceptions.ChunkedEncodingError:
        logger.warning("ChunkedEncodingError for %s. Backing off 0.5s.", url)
        time.sleep(0.5)
        try:
            resp = SESSION.get(url, params=params, timeout=DEFAULT_TIMEOUT)
        except Exception as e:
            logger.error("Repeated chunk error on %s: %s", url, e)
            return None
    except Exception as e:
        logger.error("HTTP error: %s", e)
        return None

    if resp.status_code == 429:
        _sleep_if_rate_limited(resp)
        resp = SESSION.get(url, params=params, timeout=DEFAULT_TIMEOUT)

    if resp.status_code != 200:
        logger.error("GET %s failed (%s): %s", url, resp.status_code, resp.text[:500])
        return None

    try:
        return resp.json()
    except Exception as e:
        logger.error("JSON parse error for %s: %s", url, e)
        return None


def _parse_utc(dt_str: str) -> Optional[datetime]:
    if not dt_str:
        return None
    try:
        dt = parser.isoparse(dt_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None


# ---------------------- GitLab API Accessors ---------------------- #
def get_projects_in_group(group_id: str) -> List[Dict]:
    projects: List[Dict] = []
    page = 1
    logger.info("Fetching projects for group %s ...", group_id)
    while True:
        url = f"{BASE_URL}/groups/{group_id}/projects"
        params = {"per_page": DEFAULT_PER_PAGE, "page": page, "include_subgroups": True, "archived": False}
        data = _get_json(url, params)
        if data is None:
            break
        if not data:
            break
        projects.extend(data)
        page += 1
    logger.info("Group %s: %d projects", group_id, len(projects))
    return projects


def iter_project_vulnerabilities(project_id: str, state: Optional[str] = None):
    page = 1
    while True:
        url = f"{BASE_URL}/projects/{project_id}/vulnerabilities"
        params = {"per_page": DEFAULT_PER_PAGE, "page": page}
        if state:
            params["state"] = state
        data = _get_json(url, params)
        if data is None:
            break
        if not data:
            break
        for v in data:
            yield v
        page += 1


def get_open_counts(project_id: str) -> Tuple[int, int]:
    critical = 0
    high = 0
    for v in iter_project_vulnerabilities(project_id, state="detected"):
        sev = (v.get("severity") or "").lower()
        sta = (v.get("state") or "").lower()
        if sev == "critical" and sta == "detected":
            critical += 1
        elif sev == "high" and sta == "detected":
            high += 1
    return critical, high


def get_vulns_last_n_days_all_states(project_id: str, days_back: int) -> List[Dict]:
    cutoff = datetime.now(timezone.utc) - timedelta(days=days_back)
    results: List[Dict] = []
    for v in iter_project_vulnerabilities(project_id, state=None):
        created_at = _parse_utc(v.get("created_at", ""))
        if created_at and created_at >= cutoff:
            results.append(v)
    return results


def get_window_counts(project_id: str, days_back: int = 30) -> Dict[str, int]:
    items = get_vulns_last_n_days_all_states(project_id, days_back=days_back)
    critical_n = 0
    high_n = 0
    for v in items:
        sev = (v.get("severity") or "").lower()
        if sev == "critical":
            critical_n += 1
        elif sev == "high":
            high_n += 1
    return {"critical": critical_n, "high": high_n}


# ---------------------- Excel Utilities ---------------------- #
def autosize_and_style_sheet(ws):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Header (first row)
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = header_font

    # Body borders
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    # Autosize columns
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        try:
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 60)
        except Exception:
            # ignore if column_letter is not available
            pass


def add_summary_chart(ws):
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 3:
        return
    data = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Vulnerabilities Summary"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Scrum"
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 45
    chart.height = 18
    ws.add_chart(chart, "H2")


def move_sheet_first(wb, sheet_name: str):
    ws = wb[sheet_name]
    try:
        wb.move_sheet(ws, offset=-wb.index(ws))
    except Exception:
        idx = wb.sheetnames.index(sheet_name)
        wb._sheets.insert(0, wb._sheets.pop(idx))


# ---------------------- Percentage change table (30/60/90 vs current) ---------------------- #
def _compute_pct_change_against_current(window_count: int, current_count: int) -> Optional[float]:
    """
    Compute fractional percent change: (window_count - current_count) / current_count
    Returns:
      - float fraction (e.g., 0.25 for +25%) when defined
      - 0.0 when both are zero
      - None when current_count == 0 and window_count > 0 (undefined/infinite)
    """
    try:
        wc = int(window_count)
        cc = int(current_count)
    except Exception:
        return None

    if cc == 0:
        if wc == 0:
            return 0.0
        else:
            return None  # undefined / infinite -> will display '-' (GitLab-style)
    return (wc - cc) / cc


def add_percentage_change_table(ws, summary_df: pd.DataFrame):
    pct_rows = []
    for r in summary_df.to_dict("records"):
        scrum = r.get("ScrumName")
        current_c = int(r.get("Critical", 0))
        current_h = int(r.get("High", 0))

        c30 = int(r.get("30DaysCritical", 0))
        h30 = int(r.get("30DaysHigh", 0))
        c60 = int(r.get("60DaysCritical", 0))
        h60 = int(r.get("60DaysHigh", 0))
        c90 = int(r.get("90DaysCritical", 0))
        h90 = int(r.get("90DaysHigh", 0))

        pct_rows.append({
            "ScrumName": scrum,
            "30Day Critical % change": _compute_pct_change_against_current(c30, current_c),
            "30Day High % change": _compute_pct_change_against_current(h30, current_h),
            "60Day Critical % change": _compute_pct_change_against_current(c60, current_c),
            "60Day High % change": _compute_pct_change_against_current(h60, current_h),
            "90Day Critical % change": _compute_pct_change_against_current(c90, current_c),
            "90Day High % change": _compute_pct_change_against_current(h90, current_h),
        })

    # Place table two rows below the end of existing Summary table
    start_row = ws.max_row + 2

    # Title row
    title_cell = ws.cell(row=start_row, column=1, value="Percent change of CURRENT open vs last N days (positive = increase)")
    title_cell.font = Font(bold=True)
    start_row += 1

    headers = [
        "ScrumName",
        "30Day Critical % change",
        "30Day High % change",
        "60Day Critical % change",
        "60Day High % change",
        "90Day Critical % change",
        "90Day High % change",
    ]

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Header row
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.border = thin_border
        c.fill = header_fill
        c.font = header_font

    # Data rows
    for i, r in enumerate(pct_rows, start=1):
        base_row = start_row + i
        # ScrumName
        c = ws.cell(row=base_row, column=1, value=r["ScrumName"])
        c.border = thin_border

        # Percentage columns
        for j, key in enumerate(headers[1:], start=2):
            val = r.get(key)
            cell = ws.cell(row=base_row, column=j)
            cell.border = thin_border
            if val is None:
                # GitLab style: show a dash when undefined/infinite
                cell.value = "-"
            else:
                # Write fractional value; use format that includes + sign for positive
                cell.value = float(val)
                cell.number_format = "+0.00%;-0.00%;0.00%"

    # Note row
    note_row = start_row + len(pct_rows) + 2
    note_cell = ws.cell(row=note_row, column=1, value="Note: '-' appears when current open count is 0 and window count > 0.")
    note_cell.font = Font(italic=True)


# ---------------------- Main ---------------------- #
def main():
    # Sanity check samples (non-fatal)
    try:
        sample30 = get_window_counts("27224759", days_back=30)
        logger.info("Sample project sample counts fetched.")
    except Exception:
        pass

    # Collect data per scrum and write to Excel
    summary_rows: List[Dict] = []
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for scrum_name, group_id in scrum_ids.items():
            logger.info("Processing scrum '%s' (group id %s)", scrum_name.lower(), group_id)
            data_rows: List[Dict] = []

            projects = get_projects_in_group(group_id)
            for proj in projects:
                pid = proj.get("id")
                pname = proj.get("name") or str(pid)

                # Current open counts
                critical_open, high_open = get_open_counts(pid)

                # Window counts
                last30 = get_window_counts(pid, days_back=30)
                last60 = get_window_counts(pid, days_back=60)
                last90 = get_window_counts(pid, days_back=90)

                data_rows.append({
                    "ScrumName": scrum_name.lower(),
                    "ProjectName": pname.lower(),
                    "Critical": critical_open,
                    "High": high_open,
                    "30DaysCritical": last30["critical"],
                    "30DaysHigh": last30["high"],
                    "60DaysCritical": last60["critical"],
                    "60DaysHigh": last60["high"],
                    "90DaysCritical": last90["critical"],
                    "90DaysHigh": last90["high"],
                })

            # Columns & write per-scrum sheet
            columns = [
                "ScrumName", "ProjectName",
                "Critical", "High",
                "30DaysCritical", "30DaysHigh",
                "60DaysCritical", "60DaysHigh",
                "90DaysCritical", "90DaysHigh",
            ]
            df = pd.DataFrame(data_rows, columns=columns)
            sheet_name = scrum_name.lower()
            # Excel sheet name length limit: 31 chars; ensure unique trimmed name
            safe_sheet_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_sheet_name, index=False)

            # Summarize totals
            def _sum(col: str) -> int:
                return int(df[col].fillna(0).sum()) if (not df.empty and col in df.columns) else 0

            summary_rows.append({
                "ScrumName": scrum_name.lower(),
                "Critical": _sum("Critical"),
                "High": _sum("High"),
                "30DaysCritical": _sum("30DaysCritical"),
                "30DaysHigh": _sum("30DaysHigh"),
                "60DaysCritical": _sum("60DaysCritical"),
                "60DaysHigh": _sum("60DaysHigh"),
                "90DaysCritical": _sum("90DaysCritical"),
                "90DaysHigh": _sum("90DaysHigh"),
            })

        # Write Summary sheet
        summary_cols = [
            "ScrumName",
            "Critical", "High",
            "30DaysCritical", "30DaysHigh",
            "60DaysCritical", "60DaysHigh",
            "90DaysCritical", "90DaysHigh",
        ]
        summary_df = pd.DataFrame(summary_rows, columns=summary_cols)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Post-process workbook with openpyxl
    wb = load_workbook(OUTPUT_XLSX)

    # Append percent-change table with GitLab style '-'
    if "Summary" in wb.sheetnames:
        add_percentage_change_table(wb["Summary"], summary_df)

    # Style & autosize sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        autosize_and_style_sheet(ws)

    # Move Summary to first position
    if "Summary" in wb.sheetnames:
        move_sheet_first(wb, "Summary")
        # Optional: add summary chart
        # add_summary_chart(wb["Summary"])

    wb.save(OUTPUT_XLSX)
    logger.info("Report generated: %s", OUTPUT_XLSX)


if __name__ == "__main__":
    main()
